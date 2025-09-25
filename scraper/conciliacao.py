import pymupdf    
import re
import sqlite3
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, List, Tuple
from config.settings import Settings
from config.logger import configure_logger

# Configurar logger
logger = configure_logger()

class Conciliacao:
    def __init__(self):
        self.settings = Settings()
        self.DB_PATH = self.settings.DATA_DIR / "database.db"
        self.DOWNLOADS_DIR = self.settings.DOWNLOADS_DIR
        self._inicializar_banco()

    def _inicializar_banco(self):
        conn = sqlite3.connect(self.DB_PATH)
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS resultados_conciliacao (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome_banco TEXT NOT NULL,
                banco TEXT,
                agencia TEXT,
                conta TEXT,
                saldo_inicial REAL,
                saldo_atual REAL,
                diferenca REAL,
                status TEXT,
                data_processamento TEXT
            )
        """)
        conn.commit()
        conn.close()

    def _formatar_moeda(self, valor: Optional[float]) -> Optional[str]:
        """Formata valores em moeda real R$"""
        if valor is None:
            return None
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _normalizar_numero(self, raw: Optional[str]) -> Optional[float]:
        if raw is None:
            return None
        s = raw.strip().replace("R$", "").replace(" ", "").replace("\u00a0", "")
        s_filtrado = re.sub(r"[^0-9\.,]", "", s)
        if not re.search(r"\d", s_filtrado):
            return None
        if "," in s_filtrado:
            s_num = s_filtrado.replace(".", "").replace(",", ".")
        else:
            s_num = s_filtrado
        try:
            return float(s_num)
        except ValueError:
            return None

    def _extrair_valor_apos_rotulo(self, texto: str, rotulo: str) -> Tuple[Optional[str], List[str]]:
        linhas = texto.splitlines()
        candidatos: List[str] = []
        for i, linha in enumerate(linhas):
            linha_limpa = re.sub(r'^[^A-Za-z]*', '', linha.strip(), flags=re.IGNORECASE)
            if rotulo in linha_limpa.upper():
                for j in range(i+1, min(len(linhas), i+7)):
                    val_line = linhas[j].strip()
                    if not val_line or "/" in val_line:
                        continue
                    m = re.search(r'(\d[\d\.]*,\d{2})', val_line)
                    if m:
                        candidatos.append(m.group(1))
                        break
        return (candidatos[-1] if candidatos else None, candidatos)

    def _processar_pdf(self, arquivo_pdf: Path, nome_banco: str, banco: str, agencia: str, conta: str):
        try:
            # Verificar se o banco é inválido (arquivo não existe ou é vazio)
            if not arquivo_pdf.exists() or arquivo_pdf.stat().st_size == 0:
                status = "invalido"
                saldo_inicial = saldo_atual = diferenca = None
                self._salvar_resultado(nome_banco, banco, agencia, conta,
                                    saldo_inicial, saldo_atual, diferenca, status)
                return status, saldo_inicial, saldo_atual, diferenca
                
            doc = pymupdf.open(arquivo_pdf)
            if doc.page_count == 0:
                status = "invalido"  # Alterado de "pdf_vazio" para "invalido"
                saldo_inicial = saldo_atual = diferenca = None
            else:
                pagina = doc.load_page(0)
                texto = pagina.get_text("text")

                saldo_inicial_str, _ = self._extrair_valor_apos_rotulo(texto, "SALDO INICIAL")
                saldo_atual_str, _ = self._extrair_valor_apos_rotulo(texto, "SALDO ATUAL")

                saldo_inicial = self._normalizar_numero(saldo_inicial_str)
                saldo_atual = self._normalizar_numero(saldo_atual_str)

                if saldo_inicial is not None and saldo_atual is not None:
                    if saldo_inicial == saldo_atual:
                        status = "conciliar"
                        diferenca = 0.0
                    else:
                        status = "diferenca"
                        diferenca = saldo_atual - saldo_inicial
                else:
                    status = "erro_extracao"
                    diferenca = None

            self._salvar_resultado(nome_banco, banco, agencia, conta,
                                saldo_inicial, saldo_atual, diferenca, status)
            return status, saldo_inicial, saldo_atual, diferenca

        except Exception as e:
            logger.error(f"Erro ao processar {arquivo_pdf}: {e}")
            # Verificar se é um banco inválido
            if "invalido" in str(e).lower() or not arquivo_pdf.exists():
                status = "invalido"
            else:
                status = "erro_processamento"
                
            self._salvar_resultado(nome_banco, banco, agencia, conta,
                                None, None, None, status)
            return status, None, None, None

    def _salvar_resultado(self, nome_banco, banco, agencia, conta,
                        saldo_inicial, saldo_atual, diferenca, status):
        conn = sqlite3.connect(self.DB_PATH)
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO resultados_conciliacao 
            (nome_banco, banco, agencia, conta, saldo_inicial, saldo_atual, diferenca, status, data_processamento)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            nome_banco, banco, agencia, conta,
            saldo_inicial, saldo_atual, diferenca,
            status, datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        conn.commit()
        conn.close()

    def registrar_banco_invalido(self, nome_banco: str, banco: str, agencia: str, conta: str):
        logger.info(f"Registrando banco inválido: {nome_banco} - {banco}, ag {agencia}, conta {conta}")
        self._salvar_resultado(nome_banco, banco, agencia, conta,
                            None, None, None, "invalido")

    def _gerar_planilha_resultados(self):
        """Gera planilha XLSX com os resultados do banco de dados - APENAS EXECUÇÃO ATUAL"""
        try:
            conn = sqlite3.connect(self.DB_PATH)
            
            # Consulta para obter apenas os resultados da execução atual (últimos 10 minutos)
            timestamp_limite = (datetime.now() - timedelta(minutes=10)).strftime("%Y-%m-%d %H:%M:%S")
            
            query = """
            SELECT 
                nome_banco,
                banco,
                agencia,
                conta,
                CASE WHEN saldo_inicial IS NOT NULL 
                    THEN 'R$ ' || replace(printf('%.2f', saldo_inicial), '.', ',') 
                    ELSE 'N/A' END as saldo_inicial,
                CASE WHEN saldo_atual IS NOT NULL 
                    THEN 'R$ ' || replace(printf('%.2f', saldo_atual), '.', ',') 
                    ELSE 'N/A' END as saldo_atual,
                CASE WHEN diferenca IS NOT NULL 
                    THEN 'R$ ' || replace(printf('%.2f', diferenca), '.', ',') 
                    ELSE 'N/A' END as diferenca,
                CASE 
                    WHEN status = 'invalido' THEN 'Banco Inválido'
                    WHEN status = 'conciliar' THEN 'Conciliar'
                    WHEN status = 'diferenca' THEN 'Diferença'
                    WHEN status = 'erro_extracao' THEN 'Erro Extração'
                    WHEN status = 'erro_processamento' THEN 'Erro Processamento'
                    WHEN status = 'sem_arquivo' THEN 'Sem Arquivo'
                    ELSE status 
                END as status,
                data_processamento
            FROM resultados_conciliacao 
            WHERE data_processamento > ?
            ORDER BY data_processamento DESC
            """
            
            df = pd.read_sql_query(query, conn, params=(timestamp_limite,))
            conn.close()
            
            if df.empty:
                logger.warning("Nenhum resultado encontrado para a execução atual")
                # Criar DataFrame vazio com as colunas corretas
                df = pd.DataFrame(columns=[
                    'nome_banco', 'banco', 'agencia', 'conta', 
                    'saldo_inicial', 'saldo_atual', 'diferenca', 'status', 'data_processamento'
                ])
            
            # Gerar nome do arquivo com timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"resultados_conciliacao_{timestamp}.xlsx"
            caminho_arquivo = self.settings.RESULTS_DIR / nome_arquivo
            
            # Salvar como Excel com formatação profissional
            with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Resultados Conciliação', index=False)
                
                worksheet = writer.sheets['Resultados Conciliação']
                
                # Ajustar largura das colunas para layout de extrato
                column_widths = {
                    'A': 25,  # nome_banco
                    'B': 15,  # banco
                    'C': 15,  # agencia
                    'D': 20,  # conta
                    'E': 20,  # saldo_inicial
                    'F': 20,  # saldo_atual
                    'G': 20,  # diferenca
                    'H': 20,  # status
                    'I': 20   # data_processamento
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Estilizar cabeçalho profissional
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Definir bordas
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Cabeçalho azul com texto branco
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Aplicar estilo ao cabeçalho (linha 1)
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Aplicar bordas e formatação a todas as células de dados
                for row in range(2, len(df) + 2):  # Começa na linha 2 (após cabeçalho)
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
                        
                        # Formatação específica por tipo de coluna
                        if col in [5, 6, 7]:  # Colunas monetárias (E, F, G)
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.alignment = Alignment(horizontal="left")
                        
                        # Colorir células de status
                        if col == 8:  # Coluna status (H)
                            status_value = cell.value
                            if status_value == 'Banco Inválido':
                                cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Vermelho claro
                                cell.font = Font(color="FF0000", bold=True)  # Texto vermelho
                            elif status_value == 'Conciliar':
                                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
                            elif status_value == 'Diferença':
                                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Amarelo claro
                            elif 'Erro' in str(status_value):
                                cell.fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")  # Laranja claro
                
                # Congelar painel (cabeçalho fixo) se houver dados
                if len(df) > 0:
                    worksheet.freeze_panes = worksheet['A2']
                
                # Adicionar filtros se houver dados
                if len(df) > 0:
                    worksheet.auto_filter.ref = worksheet.dimensions
                
            logger.info(f"Planilha gerada: {caminho_arquivo}")
            return caminho_arquivo
            
        except Exception as e:
            logger.error(f"Erro ao gerar planilha: {e}")
            return None
    
    def execucao(self, bancos_info: list):
        logger.info("Iniciando conciliação dos extratos...")
        resultados = []
        
        for banco in bancos_info:
            nome_banco = banco["nome"]
            if banco.get("invalido"):
                self.registrar_banco_invalido(
                    nome_banco, 
                    banco["do_banco"], 
                    banco["da_agencia"], 
                    banco["da_conta"]
                )
                resultados.append({"nome": nome_banco, "status": "invalido"})
                continue

            arquivo = banco.get("arquivo")
            if arquivo and Path(arquivo).exists():
                status, si, sa, dif = self._processar_pdf(
                    Path(arquivo),
                    nome_banco,
                    banco["do_banco"], 
                    banco["da_agencia"], 
                    banco["da_conta"]
                )
                resultados.append({
                    "nome": nome_banco, 
                    "status": status, 
                    "saldo_inicial": si,
                    "saldo_atual": sa, 
                    "diferenca": dif
                })
            else:
                # Se o arquivo não existe, registrar como inválido
                self.registrar_banco_invalido(
                    nome_banco,
                    banco["do_banco"],
                    banco["da_agencia"],
                    banco["da_conta"]
                )
                resultados.append({"nome": nome_banco, "status": "invalido"})

        # Gerar planilha com resultados
        planilha_path = self._gerar_planilha_resultados()
        
        logger.info("Conciliação finalizada.")
        return {
            "status": "success", 
            "resultados": resultados,
            "planilha_gerada": str(planilha_path) if planilha_path else None
        }