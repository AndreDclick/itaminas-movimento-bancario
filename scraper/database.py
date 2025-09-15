"""
Módulo de gerenciamento de banco de dados para conciliação contábil.
Responsável por importar, processar e exportar dados de diferentes fontes
para realizar a conciliação entre sistemas financeiros e contábeis.
"""

import sqlite3
from pathlib import Path
from config.settings import Settings
from config.logger import configure_logger
from .exceptions import (
    PlanilhaFormatacaoErradaError,
    InvalidDataFormat,
    ResultsSaveError,
    ExcecaoNaoMapeadaError
)
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from difflib import get_close_matches
from workalendar.america import Brazil
from datetime import datetime, timedelta
import pandas as pd
import xml.etree.ElementTree as ET
import numpy as np
import openpyxl
import re

# Configura o logger para registrar eventos
logger = configure_logger()

class DatabaseManager:
    """
    Gerenciador de banco de dados para conciliação contábil.
    Implementa padrão Singleton para garantir apenas uma instância.
    """
    
    # Implementação do padrão Singleton
    _instance = None
    
    def __new__(cls):
        """
        Implementa o padrão Singleton para garantir apenas uma instância.
        
        Returns:
            DatabaseManager: Instância única da classe
        """
        if cls._instance is None:
            cls._instance = super(DatabaseManager, cls).__new__(cls)
            cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        """
        Inicializa o gerenciador de banco de dados.
        Evita múltiplas inicializações no padrão Singleton.
        """
        # Verifica se já foi inicializado
        if self._initialized:
            return
            
        self.settings = Settings()  # Carrega configurações
        self.conn = None  # Conexão com o banco
        self.logger = configure_logger()  # Logger específico da classe
        self._initialize_database()  # Inicializa o banco de dados
        self._initialized = True  # Marca como inicializado

    def _initialize_database(self):
        """
        Inicializa o banco de dados SQLite e cria as tabelas necessárias.
        
        Raises:
            Exception: Se ocorrer erro na inicialização do banco
        """
        try:
            # Conecta ao banco SQLite
            self.conn = sqlite3.connect(self.settings.DB_PATH, timeout=10)
            cursor = self.conn.cursor()
            
            # Cria tabela financeiro se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_FINANCEIRO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    fornecedor TEXT,
                    titulo TEXT,
                    parcela TEXT,
                    tipo_titulo TEXT,
                    data_emissao TEXT DEFAULT NULL,
                    data_vencimento TEXT DEFAULT NULL,
                    valor_original REAL DEFAULT 0,
                    saldo_devedor REAL DEFAULT 0,
                    situacao TEXT,
                    conta_contabil TEXT,
                    centro_custo TEXT,
                    excluido BOOLEAN DEFAULT 0,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Cria tabela modelo1 (contábil) se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_MODELO1} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    conta_contabil TEXT,
                    descricao_conta TEXT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    saldo_anterior REAL DEFAULT 0,
                    debito REAL DEFAULT 0,
                    credito REAL DEFAULT 0,
                    saldo_atual REAL DEFAULT 0,
                    tipo_fornecedor TEXT,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Cria tabela contas_itens se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_CONTAS_ITENS} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    conta_contabil TEXT,
                    descricao_item TEXT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    saldo_anterior REAL DEFAULT 0,
                    debito REAL DEFAULT 0,
                    credito REAL DEFAULT 0,
                    saldo_atual REAL DEFAULT 0,
                    item TEXT DEFAULT '',
                    quantidade REAL DEFAULT 1,
                    valor_unitario REAL DEFAULT 0,
                    valor_total REAL DEFAULT 0,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Cria tabela adiantamento se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_ADIANTAMENTO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    conta_contabil TEXT,
                    descricao_item TEXT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    saldo_anterior REAL DEFAULT 0,
                    debito REAL DEFAULT 0,
                    credito REAL DEFAULT 0,
                    saldo_atual REAL DEFAULT 0,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Cria tabela Adiantamento financeiro se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_RESULTADO_ADIANTAMENTO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    total_financeiro REAL DEFAULT 0,
                    total_contabil REAL DEFAULT 0,
                    diferenca REAL DEFAULT 0,
                    status TEXT CHECK(status IN ('Conferido', 'Divergente', 'Pendente')),
                    detalhes TEXT,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Cria tabela resultado (concatenação) se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_RESULTADO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    saldo_contabil REAL DEFAULT 0,
                    saldo_financeiro REAL DEFAULT 0,
                    diferenca REAL DEFAULT 0,
                    status TEXT CHECK(status IN ('Conferido', 'Divergente', 'Pendente')),
                    detalhes TEXT,
                    ordem_importancia INTEGER,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            
            # Função auxiliar para garantir que colunas existam nas tabelas
            def ensure_column(table, column, type_):
                """Garante que uma coluna exista na tabela especificada."""
                cursor.execute(f"PRAGMA table_info({table})")
                cols = [c[1] for c in cursor.fetchall()]
                if column not in cols:
                    cursor.execute(f"ALTER TABLE {table} ADD COLUMN {column} {type_}")
            
            # Garante que colunas importantes existam
            ensure_column(self.settings.TABLE_CONTAS_ITENS, 'codigo_fornecedor', 'TEXT')
            ensure_column(self.settings.TABLE_CONTAS_ITENS, 'descricao_fornecedor', 'TEXT')
            ensure_column(self.settings.TABLE_MODELO1, 'codigo_fornecedor', 'TEXT')
            ensure_column(self.settings.TABLE_MODELO1, 'descricao_fornecedor', 'TEXT')
            ensure_column(self.settings.TABLE_RESULTADO, 'ordem_importancia', 'INTEGER')
            
            self.conn.commit()  # Confirma as alterações
            logger.info("Banco de dados inicializado com sucesso")
            
        except Exception as e:
            error_msg = f"Erro ao inicializar banco de dados: {e}"
            logger.error(error_msg)
            raise ExcecaoNaoMapeadaError(error_msg) from e

    def aplicar_sugestoes_colunas(self, df, missing_mappings):
        """
        Aplica sugestões automáticas para mapeamento de colunas faltantes.
        
        Args:
            df: DataFrame com os dados
            missing_mappings: Lista de colunas que faltam mapeamento
            
        Returns:
            DataFrame: DataFrame com colunas renomeadas conforme sugestões
        """
        try:
            candidates = df.columns.tolist()
            lower_map = {c.lower(): c for c in candidates}  # Mapa case-insensitive

            # Mapeamento manual pré-definido para colunas comuns
            manual_mapping = {
                'Codigo-Nome do Fornecedor': 'fornecedor',
                'Prf-Numero Parcela': 'titulo', 
                'Tp': 'tipo_titulo',
                'Data de Emissao': 'data_emissao',
                'Data Emissão': 'data_emissao',
                'Data de Vencto': 'data_vencimento',
                'Data Vencimento': 'data_vencimento',
                'Valor Original': 'valor_original',
                'Tit Vencidos Valor nominal': 'saldo_devedor',
                'Natureza': 'situacao',
                'Porta- dor': 'centro_custo',
                'Codigo.1': 'codigo_fornecedor',
                'Descricao.1': 'descricao_fornecedor',
                'Conta': 'conta_contabil',
                'Descricao': 'descricao_conta',
                'Descricao': 'descricao_item',  # Para contas_itens
                'Codigo': 'conta_contabil'  # Para contas_itens e adiantamento
            }

            # Aplica mapeamento manual primeiro
            for src, dest in manual_mapping.items():
                if src in df.columns and dest in missing_mappings:
                    df.rename(columns={src: dest}, inplace=True)
                    logger.warning(f"Mapeamento manual aplicado: '{src}' → '{dest}'")
                    if dest in missing_mappings:
                        missing_mappings.remove(dest)

            # Tenta encontrar correspondências automáticas para colunas restantes
            for db_col in list(missing_mappings):
                # Busca por correspondência case-insensitive
                if db_col.lower() in lower_map:
                    match = lower_map[db_col.lower()]
                    logger.warning(f"Sugestão aplicada: '{match}' → '{db_col}' (case-insensitive match)")
                    df.rename(columns={match: db_col}, inplace=True)
                    if db_col in missing_mappings:
                        missing_mappings.remove(db_col)
                    continue

                # Busca por correspondência fuzzy (similaridade)
                similar = get_close_matches(db_col, candidates, n=1, cutoff=0.6)
                if similar:
                    match = similar[0]
                    logger.warning(f"Sugestão aplicada: '{match}' → '{db_col}'")
                    df.rename(columns={match: db_col}, inplace=True)
                    if db_col in missing_mappings:
                        missing_mappings.remove(db_col)
                    continue

                # Busca por correspondência fuzzy case-insensitive
                similar_lower = get_close_matches(db_col.lower(), list(lower_map.keys()), n=1, cutoff=0.6)
                if similar_lower:
                    match = lower_map[similar_lower[0]]
                    logger.warning(f"Sugestão aplicada: '{match}' → '{db_col}' (fuzzy case-insensitive)")
                    df.rename(columns={match: db_col}, inplace=True)
                    if db_col in missing_mappings:
                        missing_mappings.remove(db_col)

            return df
        except Exception as e:
            error_msg = f"Erro ao aplicar sugestões de colunas: {e}"
            logger.error(error_msg)
            raise PlanilhaFormatacaoErradaError(error_msg) from e

    def import_from_excel(self, file_path, table_name):
        """
        Importa dados de arquivo Excel/TXT/XML para a tabela especificada.
        
        Args:
            file_path: Caminho do arquivo a ser importado
            table_name: Nome da tabela destino
            
        Returns:
            bool: True se importação foi bem sucedida, False caso contrário
        """
        try:
            filename = Path(file_path).stem.lower()

            # Determina a tabela destino baseada no nome do arquivo
            if 'ctbr100' in filename:
                table_name = self.settings.TABLE_ADIANTAMENTO
            elif 'ctbr140' in filename:
                table_name = self.settings.TABLE_CONTAS_ITENS
            elif 'ctbr040' in filename:
                table_name = self.settings.TABLE_MODELO1
            elif 'finr150' in filename:
                table_name = self.settings.TABLE_FINANCEIRO

            ext = Path(file_path).suffix.lower()
            
            # Lê o arquivo conforme o formato
            if ext == ".xlsx":
                # Lê as primeiras linhas para diagnóstico
                df_sample = pd.read_excel(file_path, nrows=5)
                logger.info(f"Primeiras 5 linhas do arquivo {file_path}:")
                logger.info(df_sample.to_string())
                
                # Lê o arquivo completo a partir da linha 2 (header=1)
                df = pd.read_excel(file_path, header=1)

            elif ext == ".xml":
                try:
                    df = DatabaseManager.read_spreadsheetml(file_path)
                except Exception as e:
                    error_msg = f"Falha ao ler {file_path} como SpreadsheetML: {e}"
                    logger.error(error_msg)
                    raise InvalidDataFormat(error_msg, tipo_dado="XML") from e

            elif ext == ".txt":
                try:
                    df = pd.read_csv(file_path, sep=";", encoding="latin1", header=1)
                except Exception:
                    df = pd.read_csv(file_path, sep="\t", encoding="latin1", header=1)

            else:
                error_msg = f"Formato de arquivo não suportado: {ext}"
                logger.error(error_msg)
                raise InvalidDataFormat(error_msg, tipo_dado=ext)

            logger.info(f"Colunas originais em {file_path}: {df.columns.tolist()}")
            logger.info(f"Primeiras linhas dos dados:")
            logger.info(df.head().to_string())

            # Limpa caracteres especiais dos nomes das colunas
            df.columns = df.columns.str.replace(r'_x000D_\n', ' ', regex=True).str.strip()
            logger.info(f"Colunas após limpeza: {df.columns.tolist()}")

            # Aplica mapeamento de colunas
            column_mapping = self._get_column_mapping(Path(file_path))
            
            # Verifica se column_mapping é um dicionário válido
            if not isinstance(column_mapping, dict):
                logger.warning(f"Mapeamento de colunas inválido para {file_path}, usando mapeamento vazio")
                column_mapping = {}
                
            if column_mapping:  # Só aplica se houver mapeamento
                df.rename(columns=column_mapping, inplace=True)
                
            logger.info(f"Colunas após mapeamento: {df.columns.tolist()}")
            logger.info(f"Amostra dos dados após mapeamento:")
            logger.info(df.head().to_string())

            # Verifica colunas obrigatórias
            expected_columns = self.get_expected_columns(table_name)
            missing_mappings = [col for col in expected_columns if col not in df.columns]

            if missing_mappings:
                logger.warning(f"Colunas mapeadas não encontradas: {missing_mappings}")
                df = self.aplicar_sugestoes_colunas(df, missing_mappings)
                remaining_missing = [col for col in expected_columns if col not in df.columns]

                if remaining_missing:
                    # Tenta criar colunas ausentes com valores padrão
                    if 'parcela' in remaining_missing and 'titulo' in df.columns:
                        df['parcela'] = df['titulo'].astype(str).str.extract(r'(\d+)$').fillna('1')
                        logger.warning("Coluna 'parcela' criada a partir do título")
                        remaining_missing.remove('parcela')

                    if 'conta_contabil' in remaining_missing:
                        df['conta_contabil'] = 'CONTA_NAO_IDENTIFICADA'
                        logger.warning("Coluna 'conta_contabil' preenchida com valor padrão para arquivo financeiro")
                        remaining_missing.remove('conta_contabil')

                    if remaining_missing:
                        error_msg = f"Colunas obrigatórias ausentes após tratamento: {remaining_missing}"
                        logger.error(error_msg)
                        raise PlanilhaFormatacaoErradaError(error_msg, caminho_arquivo=file_path)

            # Limpa e prepara os dados - com diagnóstico detalhado para datas
            logger.info("Iniciando limpeza dos dados...")
            df = self._clean_dataframe(df, table_name.lower())
            
            # Verificação específica das colunas de data
            for date_col in ['data_emissao', 'data_vencimento']:
                if date_col in df.columns:
                    logger.info(f"Coluna {date_col} - Valores únicos: {df[date_col].unique()}")
                    logger.info(f"Coluna {date_col} - Tipos: {df[date_col].dtype}")
                    logger.info(f"Coluna {date_col} - Não nulos: {df[date_col].notna().sum()}")

            # Mantém apenas colunas que existem na tabela destino
            table_columns = [col[1] for col in self.conn.execute(f"PRAGMA table_info({table_name})").fetchall()]
            keep = [col for col in df.columns if col in table_columns]
            df = df[keep]

            for col in table_columns:
                if col not in df.columns:
                    if col == 'excluido':
                        df[col] = 0  # Valor padrão para coluna excluido
                    else:
                        df[col] = None  # Valor padrão para outras colunas

            # Insere dados no banco
            df.to_sql(table_name, self.conn, if_exists='replace', index=False)
            logger.info(f"Dados importados para '{table_name}' com sucesso.")
            return True

        except (PlanilhaFormatacaoErradaError, InvalidDataFormat) as e:
            logger.error(f"Falha ao importar {file_path}: {e}")
            return False
        except Exception as e:
            error_msg = f"Erro inesperado ao importar {file_path}: {e}"
            logger.error(error_msg, exc_info=True)
            raise ExcecaoNaoMapeadaError(error_msg) from e

    @staticmethod
    def read_spreadsheetml(path: str) -> pd.DataFrame:
        """
        Lê arquivos XML no formato SpreadsheetML.
        
        Args:
            path: Caminho do arquivo XML
            
        Returns:
            DataFrame: Dados lidos do arquivo XML
            
        Raises:
            ValueError: Se não encontrar cabeçalho e dados suficientes
        """
        try:
            ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
            tree = ET.parse(path)
            root = tree.getroot()

            rows = []
            for row in root.findall(".//ss:Row", ns):
                values = []
                for cell in row.findall("ss:Cell", ns):
                    data = cell.find("ss:Data", ns)
                    values.append(data.text if data is not None else None)
                rows.append(values)

            if not rows or len(rows) < 2:
                raise ValueError("Não foi possível encontrar cabeçalho e dados no arquivo XML")

            # Pula a primeira linha (título "Item Conta")
            header = rows[1]
            data = rows[2:]

            df = pd.DataFrame(data, columns=header)

            # Deduplicar nomes de colunas manualmente
            counts = {}
            new_columns = []
            for col in df.columns:
                if col not in counts:
                    counts[col] = 0
                    new_columns.append(col)
                else:
                    counts[col] += 1
                    new_columns.append(f"{col}_{counts[col]}")

            df.columns = new_columns
            return df
        except Exception as e:
            error_msg = f"Erro ao ler arquivo SpreadsheetML {path}: {e}"
            logger.error(error_msg)
            raise InvalidDataFormat(error_msg, tipo_dado="XML") from e

    def get_expected_columns(self, table_name):
        """
        Retorna lista de colunas esperadas para cada tipo de tabela.
        
        Args:
            table_name: Nome da tabela
            
        Returns:
            list: Lista de colunas esperadas
            
        Raises:
            ValueError: Se a tabela for desconhecida
        """
        if table_name == self.settings.TABLE_FINANCEIRO:
            return [
                'fornecedor', 'titulo', 'parcela', 'tipo_titulo',
                'data_emissao', 'data_vencimento', 'valor_original',
                'saldo_devedor', 'situacao', 'conta_contabil', 'centro_custo'
            ]
        elif table_name == self.settings.TABLE_MODELO1:
            return [
                'conta_contabil', 'descricao_conta',
                'saldo_anterior', 'debito', 'credito', 'saldo_atual'
            ]
        elif table_name == self.settings.TABLE_CONTAS_ITENS:
            return [
                'conta_contabil', 'descricao_item',
                'codigo_fornecedor', 'descricao_fornecedor',
                'saldo_anterior', 'debito', 'credito', 'saldo_atual'
            ]
        elif table_name == self.settings.TABLE_ADIANTAMENTO:
            return [
                'conta_contabil', 'descricao_item',
                'codigo_fornecedor', 'descricao_fornecedor',
                'saldo_anterior', 'debito', 'credito', 'saldo_atual'
            ]
        else:
            error_msg = f"Tabela desconhecida: {table_name}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
    def _clean_dataframe(self, df, sheet_type):
        """
        Executa limpeza geral do DataFrame baseado no tipo de planilha.
        
        Args:
            df: DataFrame a ser limpo
            sheet_type: Tipo de planilha ('financeiro', 'modelo1', 'contas_itens')
            
        Returns:
            DataFrame: DataFrame limpo
            
        Raises:
            Exception: Se ocorrer erro na limpeza
        """
        try:
            # Limpa strings e remove valores vazios
            df = df.map(lambda x: str(x).strip() if pd.notna(x) else x)
            df = df.replace(['nan', 'None', ''], np.nan)
            df = df.dropna(how='all')  # Remove linhas completamente vazias
            
            # Aplica limpeza específica por tipo de planilha
            if sheet_type == 'financeiro':
                df = self._clean_financeiro_data(df)
            elif sheet_type == 'modelo1':
                df = self._clean_modelo1_data(df)
            elif sheet_type == 'contas_itens':
                df = self._clean_contas_itens_data(df)
            
            df = df.drop_duplicates()  # Remove duplicatas
            logger.info(f"DataFrame limpo - shape final: {df.shape}")
            return df
        except Exception as e:
            error_msg = f"Erro na limpeza dos dados ({sheet_type}): {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise InvalidDataFormat(error_msg, tipo_dado=sheet_type) from e

    def _clean_financeiro_data(self, df):
        """
        Limpeza específica para dados financeiros.
        
        Args:
            df: DataFrame com dados financeiros
            
        Returns:
            DataFrame: DataFrame limpo
        """
        try:
            # Diagnóstico inicial das colunas de data
            logger.info("Iniciando limpeza de dados financeiros...")
            for date_col in ['data_emissao', 'data_vencimento']:
                if date_col in df.columns:
                    logger.info(f"Coluna {date_col} antes da limpeza:")
                    logger.info(f"  Tipo: {df[date_col].dtype}")
                    logger.info(f"  Primeiros valores: {df[date_col].head().tolist()}")
                    logger.info(f"  Valores únicos: {df[date_col].unique()[:5]}")

            for date_col in ['data_emissao', 'data_vencimento']:
                if date_col in df.columns:
                    try:
                        # Converte para string primeiro para garantir consistência
                        df[date_col] = df[date_col].astype(str)
                        
                        # Remove espaços em branco
                        df[date_col] = df[date_col].str.strip()
                        
                        # Substitui valores vazios por NaN
                        df[date_col] = df[date_col].replace(['', 'nan', 'None', 'NaT'], np.nan)
                        
                        # Tenta converter para datetime
                        df[date_col] = pd.to_datetime(
                        df[date_col],
                        dayfirst=True,   # aceita tanto 01/09/2025 quanto 2025-09-01
                        errors='coerce'
                    )

                        
                        # Se não conseguir com formato específico, tenta inferir
                        if df[date_col].isna().any():
                            df[date_col] = df[date_col].dt.strftime('%Y-%m-%d')

                        
                        # Formata para string no formato brasileiro
                        df[date_col] = df[date_col].dt.strftime('%d/%m/%Y')
                        
                        # Substitui NaT por None
                        df[date_col] = df[date_col].replace('NaT', None)
                        
                    except Exception as e:
                        logger.warning(f"Erro ao converter {date_col}: {e}")
                        df[date_col] = None
            
            # Diagnóstico após a limpeza
            for date_col in ['data_emissao', 'data_vencimento']:
                if date_col in df.columns:
                    logger.info(f"Coluna {date_col} após limpeza:")
                    logger.info(f"  Tipo: {df[date_col].dtype}")
                    logger.info(f"  Primeiros valores: {df[date_col].head().tolist()}")
                    logger.info(f"  Não nulos: {df[date_col].notna().sum()}")
            
            # Remove registros de fornecedores NDF/PA
            # if 'fornecedor' in df.columns:
            #     mask = df['fornecedor'].str.contains(r'\bNDF\b|\bPA\b', case=False, na=False)
            #     logger.info(f"Removendo {mask.sum()} registros de NDF/PA/BOL/EMP/TX/INS/ISS/TXA/IRF")
            #     df = df[~mask]
            
            # Garante que todas as colunas obrigatórias existam
            required_cols = ['fornecedor', 'titulo', 'parcela', 'tipo_titulo', 
                            'data_emissao', 'data_vencimento', 'valor_original',
                            'saldo_devedor', 'situacao', 'conta_contabil', 'centro_custo']
            
            for col in required_cols:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Limpa e converte colunas numéricas
            num_cols = ['valor_original', 'saldo_devedor', 'titulos_vencer']  # ADICIONADO titulos_vencer
            for col in num_cols:
                if col in df.columns:
                    # Converte para string primeiro
                    df[col] = df[col].astype(str)
                    
                    # 1. Mantém apenas dígitos, vírgula, ponto e sinal
                    df[col] = df[col].str.replace(r'[^\d,.-]', '', regex=True)

                    # 2. Remove pontos de milhar
                    df[col] = df[col].str.replace('.', '', regex=False)

                    # 3. Troca vírgula por ponto (decimal BR → padrão Python)
                    df[col] = df[col].str.replace(',', '.', regex=False)

                    # 4. Converte para float
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

            logger.info(f"DataFrame final - shape: {df.shape}")
            logger.info(f"Colunas finais: {df.columns.tolist()}")
            
            return df
        except Exception as e:
            error_msg = f"Erro na limpeza de dados financeiros: {e}"
            logger.error(error_msg)
            raise InvalidDataFormat(error_msg, tipo_dado="financeiro") from e

    def _clean_modelo1_data(self, df):
        """
        Limpeza específica para dados do modelo1 (ctbr040).
        
        Args:
            df: DataFrame com dados do modelo1
            
        Returns:
            DataFrame: DataFrame limpo
        """
        try:
            # Classifica tipo de fornecedor baseado na descrição da conta
            if 'descricao_conta' in df.columns:
                df['tipo_fornecedor'] = df['descricao_conta'].apply(
                    lambda x: 'FORNECEDOR NACIONAL' if 'FORNEC' in str(x).upper() and 'NAC' in str(x).upper()
                    else 'FORNECEDOR' if 'FORNEC' in str(x).upper()
                    else 'OUTROS'
                )
            
            # Preenche códigos e descrições de fornecedor
            if 'codigo_fornecedor' not in df.columns:
                df['codigo_fornecedor'] = None
            if 'descricao_fornecedor' not in df.columns:
                df['descricao_fornecedor'] = None
            
            # Tenta extrair código do fornecedor da descrição da conta
            if df['codigo_fornecedor'].isna().all() and 'descricao_conta' in df.columns:
                df['codigo_fornecedor'] = df['descricao_conta'].str.extract(r'(\d{4,})', expand=False)
                df['descricao_fornecedor'] = df['descricao_conta']

            # Limpa e converte colunas numéricas
            num_cols = ['saldo_anterior', 'debito', 'credito', 'saldo_atual']
            for col in num_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(
                        df[col].astype(str)
                        .str.replace(r'[^\d,-]', '', regex=True)
                        .str.replace(',', '.'),
                        errors='coerce'
                    ).fillna(0)
            
            return df
        except Exception as e:
            error_msg = f"Erro na limpeza de dados do modelo1: {e}"
            logger.error(error_msg)
            raise InvalidDataFormat(error_msg, tipo_dado="modelo1") from e

    def _clean_contas_itens_data(self, df):
        """
        Limpa e padroniza os dados da planilha de Contas x Itens.
        """
        try:
            # Primeiro aplica o mapeamento reverso (valores para chaves)
            reverse_mapping = {v: k for k, v in self.settings.COLUNAS_CONTAS_ITENS.items()}
            df = df.rename(columns=reverse_mapping)
            
            # Garante que todas as colunas do mapeamento existam
            for col in self.settings.COLUNAS_CONTAS_ITENS.keys():
                if col not in df.columns:
                    df[col] = None
            
            return df
        except Exception as e:
            logger.error(f"Erro ao limpar dados de Contas x Itens: {e}")
            raise

    def _clean_adiantamento_data(self, df):
        """
        Limpa e padroniza os dados da planilha de Adiantamentos.
        """
        try:
            mapping = self.settings.COLUNAS_ADIANTAMENTO
            # aplica rename
            df = df.rename(columns={v: k for k, v in mapping.items()})
            # garante que todas as colunas do mapping existem
            for col in mapping.keys():
                if col not in df.columns:
                    df[col] = None
            return df[list(mapping.keys())]
        except Exception as e:
            logger.error(f"Erro ao limpar dados de Adiantamentos: {e}")
            raise


    
    def _get_column_mapping(self, file_path: Path):
        """
        Retorna mapeamento de colunas baseado no nome do arquivo e extensão.
        
        Args:
            file_path: Caminho do arquivo
            
        Returns:
            dict: Dicionário com mapeamento de colunas
            
        Raises:
            ValueError: Se o tipo de planilha não for reconhecido
        """
        filename = file_path.stem.lower()
        ext = file_path.suffix.lower()
        
        # Mapeamento para arquivos XLSX
        if ext == ".xlsx":
            if 'finr150' in filename:
                return getattr(self.settings, 'COLUNAS_FINANCEIRO', {})
            elif 'ctbr040' in filename:
                return getattr(self.settings, 'COLUNAS_MODELO1', {})
            elif 'ctbr140' in filename:
                return getattr(self.settings, 'COLUNAS_CONTAS_ITENS', {})
            elif 'ctbr100' in filename:  
                return getattr(self.settings, 'COLUNAS_ADIANTAMENTO', {})
            
        # Diagnóstico inicial das colunas de data antes do mapeamento
        if 'Data Emissao' in df.columns or 'Data Emissão' in df.columns:
            col = 'Data Emissao' if 'Data Emissao' in df.columns else 'Data Emissão'
            logger.info(f"Valores originais da coluna {col}: {df[col].head().tolist()}")

        if 'Data Vencto' in df.columns or 'Data Vencimento' in df.columns:
            col = 'Data Vencto' if 'Data Vencto' in df.columns else 'Data Vencimento'
            logger.info(f"Valores originais da coluna {col}: {df[col].head().tolist()}")

        
        # Mapeamento para arquivos TXT
        elif ext == ".txt":
            if 'ctbr140' in filename:
                return getattr(self.settings, 'COLUNAS_CONTAS_ITENS', {})
            elif 'ctbr100' in filename:
                return getattr(self.settings, 'COLUNAS_ADIANTAMENTO', {})
        
        # Mapeamento para arquivos XML
        elif ext == ".xml":
            if 'ctbr140' in filename:
                return getattr(self.settings, 'COLUNAS_CONTAS_ITENS', {})
            elif 'ctbr100' in filename:
                return getattr(self.settings, 'COLUNAS_ADIANTAMENTO', {})
        
        # Se não encontrou mapeamento, retorna dicionário vazio
        logger.warning(f"Tipo de planilha não reconhecido: {file_path.name}")
        return {}
        
    def process_data(self):
        """
        Processa os dados importados e gera resultados da conciliação.
        
        Returns:
            bool: True se o processamento foi bem sucedido, False caso contrário
        """
        try:
            self.conn.execute("BEGIN TRANSACTION")  # Inicia transação
            cursor = self.conn.cursor()
            
            # Obtém período de referência
            data_inicial, data_final = self._get_datas_referencia()
            # Limpa tabela de resultados anterior
            cursor.execute(f"DELETE FROM {self.settings.TABLE_RESULTADO}")
            cursor.execute(f"DELETE FROM {self.settings.TABLE_RESULTADO_ADIANTAMENTO}")

            query_financeiro = f"""
                INSERT INTO {self.settings.TABLE_RESULTADO}
                (codigo_fornecedor, descricao_fornecedor, saldo_financeiro, status)
                SELECT 
                    TRIM(fornecedor) as codigo_fornecedor,
                    TRIM(fornecedor) as descricao_fornecedor,
                    SUM(COALESCE(saldo_devedor, 0)) as saldo_financeiro,
                    'Pendente' as status
                FROM 
                    {self.settings.TABLE_FINANCEIRO}
                WHERE 
                    excluido = 0
                    
                GROUP BY 
                    TRIM(fornecedor)
            """
            
            # WHERE 
            #         excluido = 0
            #         AND UPPER(tipo_titulo) NOT IN ('NDF', 'PA', 'BOL', 'EMP', 'TX', 'INS', 'ISS', 'TXA', 'IRF')
            #     GROUP BY 
            #         TRIM(fornecedor)

            # AND (data_vencimento IS NULL OR data_vencimento BETWEEN '{data_inicial}' AND '{data_final}')
            cursor.execute(query_financeiro)
            
            # Atualiza com dados contábeis do modelo1 (ctbr040)
            query_contabil_update = f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET 
                    saldo_contabil = (
                        SELECT COALESCE(SUM(saldo_atual), 0)
                        FROM {self.settings.TABLE_MODELO1} m
                        WHERE 
                            (m.codigo_fornecedor = {self.settings.TABLE_RESULTADO}.codigo_fornecedor AND m.codigo_fornecedor IS NOT NULL AND m.codigo_fornecedor <> '')
                            OR (m.descricao_fornecedor = {self.settings.TABLE_RESULTADO}.descricao_fornecedor)
                    ),
                    detalhes = (
                        SELECT GROUP_CONCAT(COALESCE(m.tipo_fornecedor, '') || ': ' || m.saldo_atual, ' | ')
                        FROM {self.settings.TABLE_MODELO1} m
                        WHERE 
                            (m.codigo_fornecedor = {self.settings.TABLE_RESULTADO}.codigo_fornecedor AND m.codigo_fornecedor IS NOT NULL AND m.codigo_fornecedor <> '')
                            OR (m.descricao_fornecedor = {self.settings.TABLE_RESULTADO}.descricao_fornecedor)
                    )
                WHERE EXISTS (
                    SELECT 1
                    FROM {self.settings.TABLE_MODELO1} m2
                    WHERE 
                        (m2.codigo_fornecedor = {self.settings.TABLE_RESULTADO}.codigo_fornecedor AND m2.codigo_fornecedor IS NOT NULL AND m2.codigo_fornecedor <> '')
                        OR (m2.descricao_fornecedor = {self.settings.TABLE_RESULTADO}.descricao_fornecedor)
                )
            """
            cursor.execute(query_contabil_update)
                        
            # Adiciona dados de adiantamentos
            query_adiantamento = f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET 
                    saldo_contabil = saldo_contabil + (
                        SELECT COALESCE(SUM(saldo_atual),0)
                        FROM {self.settings.TABLE_ADIANTAMENTO} a
                        WHERE a.codigo_fornecedor = {self.settings.TABLE_RESULTADO}.codigo_fornecedor
                    )
                WHERE EXISTS (
                    SELECT 1
                    FROM {self.settings.TABLE_ADIANTAMENTO} a2
                    WHERE a2.codigo_fornecedor = {self.settings.TABLE_RESULTADO}.codigo_fornecedor
                )
            """
            cursor.execute(query_adiantamento)
            
            # Insere dados contábeis que não tiveram match financeiro
            query_contabeis_sem_match = f"""
                INSERT INTO {self.settings.TABLE_RESULTADO}
                (codigo_fornecedor, descricao_fornecedor, saldo_contabil, status, detalhes)
                SELECT 
                    CASE 
                        WHEN UPPER(m.descricao_conta) LIKE 'FORNECEDORES%' 
                        THEN 'FORNECEDORES' 
                        ELSE COALESCE(NULLIF(TRIM(m.codigo_fornecedor), ''), 
                                    SUBSTR(m.descricao_conta, 1, INSTR(m.descricao_conta || ' ', ' ') - 1)) 
                    END as codigo_fornecedor,
                    CASE 
                        WHEN UPPER(m.descricao_conta) LIKE 'FORNECEDORES%' 
                        THEN 'FORNECEDORES (Consolidado)' 
                        ELSE COALESCE(NULLIF(TRIM(m.descricao_fornecedor), ''), m.descricao_conta) 
                    END as descricao_fornecedor,
                    SUM(m.saldo_atual) as saldo_contabil,
                    'Pendente' as status,
                    m.tipo_fornecedor as detalhes
                FROM 
                    {self.settings.TABLE_MODELO1} m
                WHERE 
                    m.descricao_conta LIKE 'FORNEC%'
                    AND NOT EXISTS (
                        SELECT 1
                        FROM {self.settings.TABLE_RESULTADO} r
                        WHERE r.codigo_fornecedor = COALESCE(NULLIF(TRIM(m.codigo_fornecedor), ''), 
                                                        SUBSTR(m.descricao_conta, 1, INSTR(m.descricao_conta || ' ', ' ') - 1))
                    )
                GROUP BY 
                    CASE 
                        WHEN UPPER(m.descricao_conta) LIKE 'FORNECEDORES%' THEN 'FORNECEDORES'
                        ELSE COALESCE(NULLIF(TRIM(m.codigo_fornecedor), ''), 
                                    SUBSTR(m.descricao_conta, 1, INSTR(m.descricao_conta || ' ', ' ') - 1))
                    END,
                    CASE 
                        WHEN UPPER(m.descricao_conta) LIKE 'FORNECEDORES%' THEN 'FORNECEDORES (Consolidado)'
                        ELSE COALESCE(NULLIF(TRIM(m.descricao_fornecedor), ''), m.descricao_conta)
                    END,
                    m.tipo_fornecedor
            """
            cursor.execute(query_contabeis_sem_match)

            # Calcula diferenças e define status
            query_diferenca = f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET 
                    diferenca = ROUND(COALESCE(saldo_financeiro,0) - COALESCE(saldo_contabil,0), 2),
                    status = CASE 
                        WHEN saldo_contabil IS NULL AND saldo_financeiro IS NULL THEN 'Pendente'
                        WHEN ABS(COALESCE(saldo_financeiro,0) - COALESCE(saldo_contabil,0)) <= 
                            (0.03 * CASE 
                                WHEN ABS(COALESCE(saldo_contabil,0)) > ABS(COALESCE(saldo_financeiro,0)) 
                                THEN ABS(COALESCE(saldo_contabil,0)) 
                                ELSE ABS(COALESCE(saldo_financeiro,0)) 
                            END)
                            THEN 'Conferido' 
                        ELSE 'Divergente' 
                    END
            """
            cursor.execute(query_diferenca)
            
            # Para fornecedores com status DIVERGENTE, busca detalhes na tabela contas_itens
            query_investigacao = f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET detalhes = (
                    SELECT 'Divergência: R$ ' || ABS(diferenca) || 
                        '. Itens Contábeis encontrados: ' || 
                        COALESCE(
                            (SELECT GROUP_CONCAT(
                                    'Item: ' || ci.descricao_item || 
                                    ' (Valor: R$ ' || ci.saldo_atual || ')', 
                                    '; '
                                )
                                FROM contas_itens ci
                                WHERE ci.conta_contabil LIKE '%' || {self.settings.TABLE_RESULTADO}.codigo_fornecedor || '%'
                                LIMIT 5),  -- Limita a 5 itens para não ficar muito longo
                            'Nenhum item específico encontrado'
                        )
                )
                WHERE status = 'Divergente'
                AND EXISTS (
                    SELECT 1 FROM contas_itens ci2 
                    WHERE ci2.conta_contabil LIKE '%' || {self.settings.TABLE_RESULTADO}.codigo_fornecedor || '%'
                )
            """
            cursor.execute(query_investigacao)
            
            # Para fornecedores divergentes sem itens específicos na contas_itens
            cursor.execute(f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET detalhes = 'Divergência: R$ ' || ABS(diferenca) || 
                            '. Investigar manualmente no sistema. Nenhum item contábil específico encontrado para análise automática.'
                WHERE status = 'Divergente' 
                AND (detalhes IS NULL OR detalhes = '' OR detalhes LIKE '%Conciliação%')
            """)
            
            # Classifica por ordem de importância (magnitude da diferença)
            try:
                cursor.execute(f"""
                    UPDATE {self.settings.TABLE_RESULTADO}
                    SET ordem_importancia = (
                        SELECT COUNT(*) 
                        FROM {self.settings.TABLE_RESULTADO} r2 
                        WHERE ABS(COALESCE(r2.diferenca,0)) >= ABS(COALESCE({self.settings.TABLE_RESULTADO}.diferenca,0))
                    )
                """)
            except Exception as rank_error:
                logger.error(f"Erro ao classificar por importância: {rank_error}")

            # Atualiza detalhes para registros não divergentes
            cursor.execute(f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET detalhes = 
                    CASE 
                        WHEN status = 'OK' THEN 'Conciliação perfeita'
                        WHEN status = 'PENDENTE' THEN 'Financeiro: ' || COALESCE(saldo_financeiro,0) || 
                                                    ' | Contábil: ' || COALESCE(saldo_contabil,0) || 
                                                    ' | Diferença: ' || COALESCE(diferenca,0)
                        ELSE detalhes  -- Mantém os detalhes da investigação para divergências
                    END
            """)

            # NOVO: Processamento específico para adiantamentos
            self._process_adiantamentos()

            self.conn.commit()  # Confirma transação
            logger.info("Processamento de dados concluído com sucesso")
            return True
            
        except Exception as e:
            error_msg = f"Erro ao processar dados: {e}"
            logger.error(error_msg, exc_info=True)
            self.conn.rollback()  # Reverte em caso de erro
            raise ExcecaoNaoMapeadaError(error_msg) from e
    

    # def _get_datas_referencia(self):
    #     """
    #     Retorna datas de referência para o processamento da conciliação.
    #     Executa nos dias 20 e último dia do mês, ajustando para dias úteis.
        
    #     Returns:
    #         tuple: (data_inicial, data_final) formatadas como strings
    #     """
    #     cal = Brazil()
    #     hoje = datetime.now().date()
        
    #     # Verifica se é dia 20 ou último dia do mês
    #     ultimo_dia_mes = self._ultimo_dia_mes(hoje)
        
    #     if hoje.day == 20 or hoje == ultimo_dia_mes:
    #         # Ajusta para o próximo dia útil se for fim de semana/feriado
    #         data_execucao = cal.add_working_days(hoje, 0)
    #     else:
    #         # Se não for dia de execução, retorna None
    #         return None, None
        
    #     # Define as datas de referência baseadas no dia de execução
    #     if data_execucao.day == 20:
    #         # Execução no dia 20 - refere-se ao mês atual
    #         data_inicial = data_execucao.replace(day=1)
    #         data_final = data_execucao.replace(day=20)
    #     else:
    #         # Execução no último dia - refere-se ao mês anterior
    #         mes_anterior = data_execucao.replace(day=1) - timedelta(days=1)
    #         data_inicial = mes_anterior.replace(day=1)
    #         data_final = self._ultimo_dia_mes(mes_anterior)
        
    #     return data_inicial.strftime("%d/%m/%Y"), data_final.strftime("%d/%m/%Y")

    def _get_datas_referencia(self):
        try:
            cal = Brazil()
            hoje = datetime.now().date()
            data_inicial = hoje.replace(day=1)
            data_inicial = cal.add_working_days(data_inicial - timedelta(days=1), 1)
            data_final = hoje
            # Retorna em ISO (YYYY-MM-DD) — adequado para BETWEEN no SQLite
            # return data_inicial.strftime("01/05/2025"), data_final.strftime("01/06/2025")
            return data_inicial.strftime("%d/%m/%Y"), data_final.strftime("%d/%m/%Y")

        except Exception as e:
            error_msg = f"Erro ao obter datas de referência: {e}"
            logger.error(error_msg)
            raise ExcecaoNaoMapeadaError(error_msg) from e


    def _ultimo_dia_mes(self, date):
        """
        Retorna o último dia do mês da data fornecida.
        
        Args:
            date: Data para calcular o último dia do mês
            
        Returns:
            datetime: Último dia do mês
        """
        next_month = date.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def validate_data_consistency(self):
        try:
            cursor = self.conn.cursor()
            
            # Verifica totais financeiros vs contábeis
            query = f"""
                SELECT 
                    (SELECT SUM(saldo_devedor) FROM {self.settings.TABLE_FINANCEIRO} 
                    WHERE excluido = 0 AND UPPER(tipo_titulo) NOT IN ('NDF', 'PA', 'BOL', 'EMP', 'TX', 'INS', 'ISS', 'TXA', 'IRF')) as total_financeiro,
                    (SELECT SUM(saldo_atual) FROM {self.settings.TABLE_MODELO1} 
                    WHERE descricao_conta LIKE 'FORNECEDOR%') as total_contabil
            """
            cursor.execute(query)
            totals = cursor.fetchone()
            
            diferenca_percentual = abs(totals[0] - totals[1]) / max(totals[0], totals[1]) * 100
            
            if diferenca_percentual > 5:  # Tolerância de 5% para o total
                logger.warning(f"Diferença significativa entre totais: Financeiro={totals[0]}, Contábil={totals[1]}")
            
            return True
        except Exception as e:
            error_msg = f"Erro na validação de consistência: {e}"
            logger.error(error_msg)
            return False

    def _apply_metadata_styles(self, worksheet):
        """
        Aplica estilos à aba de metadados de forma otimizada.
        """
        try:
            # Define estilos
            title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_font = Font(color="FFFFFF", bold=True, size=14)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Formata título (primeira linha)
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = title_fill
                    cell.font = title_font
                    cell.border = thin_border
            
            # Formata cabeçalho (segunda linha)
            for row in worksheet.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
            
            # Aplica bordas a todas as células restantes
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
            
            # Ajusta largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)  # Limita a largura máxima
                worksheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            logger.warning(f"Erro ao aplicar estilos de metadados: {e}")
    def _apply_styles(self, worksheet):
        """
        Aplica estilos visuais básicos à planilha Excel de forma otimizada.
        """
        try:
            # Define estilos
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            align_center = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Aplica estilos ao cabeçalho em lote
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = thin_border
            
            # Identifica colunas monetárias
            header = [c.value for c in worksheet[1] if c.value is not None]
            monetary_headers = {
                'Valor Financeiro', 'Valor Contábil', 'Diferença',
                'Valor em Aberto', 'Valor Provisionado', 'Saldo Atual',
                'Débito', 'Crédito', 'Valor Original', 'Saldo Devedor',
                'Quantidade', 'Valor Unitário', 'Valor Total'
            }
            
            # Aplica formatação monetária em colunas inteiras 
            for col_idx, col_name in enumerate(header, 1):
                if col_name in monetary_headers:
                    col_letter = get_column_letter(col_idx)
                    for cell in worksheet[col_letter][1:]:  # Pula o cabeçalho
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = 'R$ #,##0.00;[Red]R$ -#,##0.00'
            
            # Aplica bordas a todas as células (em lote por linha)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
            
            # Ajusta largura das colunas automaticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)  # Limita a largura máxima
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Erro ao aplicar estilos básicos: {e}")

    def _apply_enhanced_styles(self, worksheet, stats):
        """
        Aplica estilos visuais melhorados com formatação otimizada.
        """
        try:
            # Define estilos
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            align_center = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Aplica estilos ao cabeçalho em lote
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = thin_border
            
            # Identifica índices de colunas para formatação monetária
            header = [c.value for c in worksheet[1] if c.value is not None]
            monetary_columns = [
                'Valor Financeiro', 'Valor Contábil', 'Diferença',
                'Valor em Aberto', 'Valor Provisionado', 'Saldo Atual',
                'Débito', 'Crédito'
            ]
            
            # Aplica formatação monetária em colunas inteiras
            for col_idx, col_name in enumerate(header, 1):
                if col_name in monetary_columns:
                    col_letter = get_column_letter(col_idx)
                    for cell in worksheet[col_letter][1:]:  # Pula o cabeçalho
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = 'R$ #,##0.00;[Red]R$ -#,##0.00'
            
            # Identifica colunas de status e diferença
            status_idx = header.index("Status") + 1 if "Status" in header else None
            diferenca_idx = header.index("Diferença") + 1 if "Diferença" in header else None
            
            # Cores para formatação condicional
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            red_font = Font(color="9C0006", bold=True)
            
            # Aplica formatação condicional por linhas (mais eficiente)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                # Formata valores negativos em vermelho (apenas para coluna Diferença)
                if diferenca_idx:
                    diff_cell = row[diferenca_idx-1]  # -1 porque index começa em 0
                    if diff_cell.value is not None and diff_cell.value < 0:
                        diff_cell.font = red_font
                
                # Formatação baseada no status
                if status_idx:
                    status_cell = row[status_idx-1]  # -1 porque index começa em 0
                    status_value = status_cell.value if status_cell.value else ""
                    
                    fill_color = None
                    if status_value == 'Conferido':
                        fill_color = green_fill
                    elif status_value == 'Divergente':
                        fill_color = red_fill
                    elif status_value == 'Pendente':
                        fill_color = yellow_fill
                    
                    # Aplica o preenchimento apenas se necessário
                    if fill_color:
                        for cell in row:
                            cell.fill = fill_color
                
                # Aplica bordas a todas as células
                for cell in row:
                    cell.border = thin_border
            
            # Ajusta largura das colunas automaticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)  # Limita a largura máxima
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Erro ao aplicar estilos avançados: {e}")

    def _optimize_worksheet_performance(self, worksheet):
        """
        Otimiza a planilha para melhor performance de escrita.
        """
        # Desativa propriedades que tornam a escrita lenta
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.showRowColHeaders = False

    def _protect_sheets(self, workbook):
        """
        Protege todas as abas, exceto a coluna 'Observações' na aba de resumo.
        """
        try:
            from openpyxl.worksheet.protection import SheetProtection
            
            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]
                if sheet is not None:
                    # Protege a planilha inteira
                    sheet.protection = SheetProtection(
                        sheet=True, 
                        selectLockedCells=False,
                        selectUnlockedCells=False
                    )
                    
                    # Libera apenas a coluna "Observações" na aba de resumo
                    if sheetname == 'Resumo da Conciliação':
                        header = [cell.value for cell in sheet[1] if cell.value is not None]
                        if "Observações" in header:
                            obs_col_idx = header.index("Observações") + 1
                            for row in range(2, sheet.max_row + 1):
                                cell = sheet.cell(row=row, column=obs_col_idx)
                                cell.protection = Protection(locked=False)
        
        except Exception as e:
            logger.warning(f"")

    def _process_adiantamentos(self):
        """Processa especificamente a conciliação de adiantamentos"""
        try:
            cursor = self.conn.cursor()
            
            # Verifica se a tabela existe e tem as colunas corretas
            cursor.execute(f"PRAGMA table_info({self.settings.TABLE_RESULTADO_ADIANTAMENTO})")
            columns = [col[1] for col in cursor.fetchall()]
            
            # Se não tem as colunas necessárias, recria a tabela
            if 'total_financeiro' not in columns:
                self._recreate_adiantamento_table()
            
            # Calcula totais financeiros de adiantamentos (NDF e PA)
            query_adiantamento_financeiro = f"""
                INSERT INTO {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                (codigo_fornecedor, descricao_fornecedor, total_financeiro, status)
                SELECT 
                    TRIM(fornecedor) as codigo_fornecedor,
                    TRIM(fornecedor) as descricao_fornecedor,
                    SUM(COALESCE(saldo_devedor, 0)) as total_financeiro, 
                    'Pendente' as status
                FROM 
                    {self.settings.TABLE_FINANCEIRO}
                WHERE 
                    excluido = 0
                    AND UPPER(tipo_titulo) IN ('NDF', 'PA')
                GROUP BY 
                    TRIM(fornecedor)
            """
            cursor.execute(query_adiantamento_financeiro)
            
            # Soma de Títulos Vencidos + Títulos a Vencer (J + K)
            query_soma_valores = f"""
                UPDATE {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                SET total_financeiro = (
                    SELECT COALESCE(SUM(
                        COALESCE(saldo_devedor, 0) + COALESCE("Titulos a vencer Valor nominal", 0)  
                    ), 0)
                    FROM {self.settings.TABLE_FINANCEIRO} f
                    WHERE f.fornecedor = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.codigo_fornecedor
                    AND f.excluido = 0
                    AND UPPER(f.tipo_titulo) IN ('NDF', 'PA')
                )
                WHERE codigo_fornecedor IN (
                    SELECT DISTINCT fornecedor 
                    FROM {self.settings.TABLE_FINANCEIRO} 
                    WHERE UPPER(tipo_titulo) IN ('NDF', 'PA')
                )
            """
            cursor.execute(query_soma_valores)
            
            # Atualiza com dados contábeis de adiantamento
            query_contabil_update = f"""
                UPDATE {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                SET 
                    total_contabil = (
                        SELECT COALESCE(SUM(saldo_atual), 0)
                        FROM {self.settings.TABLE_ADIANTAMENTO} a
                        WHERE a.codigo_fornecedor = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.codigo_fornecedor
                    ),
                    detalhes = 'Adiantamento: ' || COALESCE((
                        SELECT GROUP_CONCAT(descricao_fornecedor || ': R$ ' || saldo_atual, ' | ')
                        FROM {self.settings.TABLE_ADIANTAMENTO} a2
                        WHERE a2.codigo_fornecedor = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.codigo_fornecedor
                    ), 'Nenhum registro contábil')
                WHERE EXISTS (
                    SELECT 1
                    FROM {self.settings.TABLE_ADIANTAMENTO} a3
                    WHERE a3.codigo_fornecedor = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.codigo_fornecedor
                )
            """
            cursor.execute(query_contabil_update)
            
            # Insere adiantamentos contábeis que não tiveram match financeiro
            query_contabeis_sem_match = f"""
                INSERT INTO {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                (codigo_fornecedor, descricao_fornecedor, total_contabil, status, detalhes)
                SELECT 
                    codigo_fornecedor,
                    descricao_fornecedor,
                    SUM(saldo_atual) as total_contabil,
                    'Pendente' as status,
                    'Adiantamento contábil sem correspondência financeira' as detalhes
                FROM 
                    {self.settings.TABLE_ADIANTAMENTO}
                WHERE 
                    codigo_fornecedor IS NOT NULL 
                    AND codigo_fornecedor <> ''
                    AND NOT EXISTS (
                        SELECT 1
                        FROM {self.settings.TABLE_RESULTADO_ADIANTAMENTO} r
                        WHERE r.codigo_fornecedor = {self.settings.TABLE_ADIANTAMENTO}.codigo_fornecedor
                    )
                GROUP BY 
                    codigo_fornecedor, descricao_fornecedor
            """
            cursor.execute(query_contabeis_sem_match)
            
            # Calcula diferenças e define status
            query_diferenca = f"""
                UPDATE {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                SET 
                    diferenca = ROUND(COALESCE(total_financeiro, 0) - COALESCE(total_contabil, 0), 2),
                    status = CASE 
                        WHEN total_contabil IS NULL AND total_financeiro IS NULL THEN 'Pendente'
                        WHEN ABS(COALESCE(total_financeiro, 0) - COALESCE(total_contabil, 0)) <= 
                            (0.03 * CASE 
                                WHEN ABS(COALESCE(total_contabil, 0)) > ABS(COALESCE(total_financeiro, 0)) 
                                THEN ABS(COALESCE(total_contabil, 0)) 
                                ELSE ABS(COALESCE(total_financeiro, 0)) 
                            END)
                            THEN 'Conferido' 
                        ELSE 'Divergente' 
                    END
            """
            cursor.execute(query_diferenca)
            
        except Exception as e:
            error_msg = f"Erro no processamento de adiantamentos: {e}"
            logger.error(error_msg)
            raise

    def _recreate_adiantamento_table(self):
        """Recria a tabela de resultado_adiantamento com estrutura correta"""
        try:
            cursor = self.conn.cursor()
            cursor.execute(f"DROP TABLE IF EXISTS {self.settings.TABLE_RESULTADO_ADIANTAMENTO}")
            
            cursor.execute(f"""
                CREATE TABLE {self.settings.TABLE_RESULTADO_ADIANTAMENTO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    total_financeiro REAL DEFAULT 0,
                    total_contabil REAL DEFAULT 0,
                    diferenca REAL DEFAULT 0,
                    status TEXT CHECK(status IN ('Conferido', 'Divergente', 'Pendente')),
                    detalhes TEXT,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            self.conn.commit()
            logger.info("Tabela resultado_adiantamento recriada com estrutura correta")
        except Exception as e:
            error_msg = f"Erro ao recriar tabela resultado_adiantamento: {e}"
            logger.error(error_msg)
            raise

    def export_to_excel(self):
        """
        Exporta resultados para arquivo Excel formatado com metadados.
        """
        
        data_inicial_iso, data_final_iso = self._get_datas_referencia()
        try:
            data_inicial = datetime.strptime(data_inicial_iso, '%Y-%m-%d').strftime('%d/%m/%Y')
            data_final = datetime.strptime(data_final_iso, '%Y-%m-%d').strftime('%d/%m/%Y')
        except Exception:
            data_inicial, data_final = data_inicial_iso, data_final_iso

        output_path = self.settings.RESULTS_DIR / f"CONCILIACAO_{data_inicial.replace('/', '-')}_a_{data_final.replace('/', '-')}.xlsx"

        try:
            if not self.conn:
                error_msg = "Tentativa de exportação com conexão fechada"
                logger.error(error_msg)
                raise ResultsSaveError(error_msg, caminho=output_path)
            
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            
            # Query para obter estatísticas de processamento
            query_stats = f"""
                SELECT 
                    COUNT(*) as total_registros,
                    SUM(CASE WHEN status = 'Conferido' THEN 1 ELSE 0 END) as conciliados_ok,
                    SUM(CASE WHEN status = 'Divergente' THEN 1 ELSE 0 END) as divergentes,
                    SUM(CASE WHEN status = 'Pendente' THEN 1 ELSE 0 END) as pendentes,
                    SUM(saldo_financeiro) as total_financeiro,
                    SUM(saldo_contabil) as total_contabil,
                    (SUM(saldo_financeiro) - SUM(saldo_contabil)) as diferenca_geral,
                    SUM(CASE WHEN status = 'Divergente' THEN diferenca ELSE 0 END) as total_divergencia
                FROM 
                    {self.settings.TABLE_RESULTADO}
            """

            stats = pd.read_sql(query_stats, self.conn).iloc[0]

            # ABA: "Títulos a Pagar" (Dados Financeiros) - VERIFICAÇÃO DE DATAS
            query_financeiro = f"""
                SELECT 
                    fornecedor as "Fornecedor",
                    titulo as "Título",
                    parcela as "Parcela",
                    tipo_titulo as "Tipo Título",
                    CASE 
                        WHEN data_emissao IS NULL OR data_emissao = '' THEN NULL
                        ELSE data_emissao 
                    END as "Data Emissão",
                    CASE 
                        WHEN data_vencimento IS NULL OR data_vencimento = '' THEN NULL
                        ELSE data_vencimento 
                    END as "Data Vencimento",
                    valor_original as "Valor Original",
                    saldo_devedor as "Saldo Devedor",
                    situacao as "Situação",
                    conta_contabil as "Conta Contábil",
                    centro_custo as "Centro Custo"
                FROM 
                    {self.settings.TABLE_FINANCEIRO}
                WHERE 
                    excluido = 0
                    AND UPPER(tipo_titulo) NOT IN ('NDF', 'PA', 'BOL', 'EMP', 'TX', 'INS', 'ISS', 'TXA', 'IRF')
                ORDER BY 
                    fornecedor, titulo, parcela
            """
            df_financeiro = pd.read_sql(query_financeiro, self.conn)
            
            # Verifica se há problemas com as datas
            logger.info(f"Total de registros financeiros: {len(df_financeiro)}")
            df_financeiro['Data Emissão'] = pd.to_datetime(df_financeiro['Data Emissão'], errors='coerce').dt.strftime('%d/%m/%Y')
            df_financeiro['Data Vencimento'] = pd.to_datetime(df_financeiro['Data Vencimento'], errors='coerce').dt.strftime('%d/%m/%Y')

            
            df_financeiro.to_excel(writer, sheet_name='Títulos a Pagar', index=False)

            # NOVA ABA: "Adiantamentos de Títulos a Pagar" (Dados Financeiros) - VERIFICAÇÃO DE DATAS
            query_adi_financeiro = f"""
                SELECT 
                    fornecedor as "Fornecedor",
                    titulo as "Título",
                    parcela as "Parcela",
                    tipo_titulo as "Tipo Título",
                    CASE 
                        WHEN data_emissao IS NULL OR data_emissao = '' THEN NULL
                        ELSE data_emissao 
                    END as "Data Emissão",
                    CASE 
                        WHEN data_vencimento IS NULL OR data_vencimento = '' THEN NULL
                        ELSE data_vencimento 
                    END as "Data Vencimento",
                    valor_original as "Valor Original",
                    saldo_devedor as "Saldo Devedor",
                    situacao as "Situação",
                    conta_contabil as "Conta Contábil",
                    centro_custo as "Centro Custo"
                FROM 
                    {self.settings.TABLE_FINANCEIRO}
                WHERE 
                    excluido = 0
                    AND UPPER(tipo_titulo) NOT IN ('NF', 'FT', 'BOL', 'EMP', 'TX', 'INS', 'ISS', 'TXA', 'IRF')
                ORDER BY 
                    fornecedor, titulo, parcela
            """
            df_adi_financeiro = pd.read_sql(query_adi_financeiro, self.conn)
            
            # Verifica se há problemas com as datas
            logger.info(f"Total de registros financeiros: {len(df_adi_financeiro)}")
            df_adi_financeiro['Data Emissão'] = pd.to_datetime(df_adi_financeiro['Data Emissão'], errors='coerce').dt.strftime('%d/%m/%Y')
            df_adi_financeiro['Data Vencimento'] = pd.to_datetime(df_adi_financeiro['Data Vencimento'], errors='coerce').dt.strftime('%d/%m/%Y')

            
            df_adi_financeiro.to_excel(writer, sheet_name='Adiantamento TAP', index=False)
            
            # ABA: "Balancete" (Dados Contábeis) - TODOS OS CAMPOS
            query_contabil = f"""
                SELECT 
                    conta_contabil as "Conta Contábil",
                    descricao_conta as "Descrição Conta",
                    codigo_fornecedor as "Código Fornecedor",
                    descricao_fornecedor as "Descrição Fornecedor",
                    saldo_anterior as "Saldo Anterior",
                    debito as "Débito",
                    credito as "Crédito",
                    saldo_atual as "Saldo Atual",
                    tipo_fornecedor as "Tipo Fornecedor"
                FROM 
                    {self.settings.TABLE_MODELO1}
                WHERE 
                    descricao_conta LIKE '%FORNEC%'
                    AND conta_contabil NOT LIKE '1.01.06.02%'  -- Exclui adiantamentos detalhados
                ORDER BY 
                    conta_contabil, codigo_fornecedor
            """

            df_contabil = pd.read_sql(query_contabil, self.conn)
            df_contabil.to_excel(writer, sheet_name='Balancete', index=False)
            
            # ABA: "Adiantamento" (Dados de Adiantamentos)
            query_adiantamento = f"""
                SELECT 
                    conta_contabil as "Conta Contábil",
                    descricao_item as "Descrição Item",
                    codigo_fornecedor as "Código Fornecedor",
                    descricao_fornecedor as "Descrição Fornecedor",
                    saldo_anterior as "Saldo Anterior",
                    saldo_atual as "Saldo Atual"
                FROM 
                    {self.settings.TABLE_ADIANTAMENTO}
                ORDER BY 
                    conta_contabil, codigo_fornecedor
            """
            df_adiantamento = pd.read_sql(query_adiantamento, self.conn)
            df_adiantamento.to_excel(writer, sheet_name='Adiantamento', index=False)
            
            # ABA: "Contas x Itens" (Detalhamento Contábil)
            query_contas_itens = f"""
                SELECT 
                    conta_contabil as "Conta Contábil",
                    descricao_item as "Descrição Item",
                    codigo_fornecedor as "Código Fornecedor",
                    descricao_fornecedor as "Descrição Fornecedor",
                    saldo_anterior as "Saldo Anterior",
                    saldo_atual as "Saldo Atual"
                FROM 
                    {self.settings.TABLE_CONTAS_ITENS}
                ORDER BY 
                    conta_contabil, codigo_fornecedor
            """
            df_contas_itens = pd.read_sql(query_contas_itens, self.conn)
            df_contas_itens.to_excel(writer, sheet_name='Contas x Itens', index=False)

            # NOVA ABA: "Resumo Adiantamentos"
            query_resumo_adiantamento = f"""
                SELECT 
                    codigo_fornecedor as "Código Fornecedor",
                    descricao_fornecedor as "Descrição Fornecedor",
                    total_financeiro as "Total Financeiro",
                    total_contabil as "Total Contábil",
                    diferenca as "Diferença",
                    status as "Status",
                    detalhes as "Detalhes"
                FROM 
                    {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                ORDER BY 
                    ABS(diferenca) DESC,
                    codigo_fornecedor
            """
            df_resumo_adiantamento = pd.read_sql(query_resumo_adiantamento, self.conn)
            df_resumo_adiantamento.to_excel(writer, sheet_name='Resumo Adiantamentos', index=False)

            # ABA: "Resumo da Conciliação" (Principal)
            query_resumo = f"""
                SELECT 
                    
                    descricao_fornecedor as "Descrição Fornecedor",
                    saldo_contabil as "Saldo Contábil",
                    saldo_financeiro as "Saldo Financeiro",
                    (saldo_contabil - saldo_financeiro) as "Diferença (Contábil - Financeiro)",
                    CASE 
                        WHEN (saldo_contabil - saldo_financeiro) > 0 THEN 'Contábil > Financeiro'
                        WHEN (saldo_contabil - saldo_financeiro) < 0 THEN 'Financeiro > Contábil'
                        ELSE 'Igual'
                    END as "Tipo Diferença",
                    CASE 
                        WHEN status = 'Conferido' THEN 'Conferido'  
                        WHEN status = 'Divergente' THEN 'Divergente'
                        ELSE 'Pendente'
                    END as "Status", 
                    detalhes as "Detalhes",
                    '{data_inicial} a {data_final}' as "Data de Referência"
                FROM 
                    {self.settings.TABLE_RESULTADO}
                ORDER BY 
                    ABS(saldo_contabil - saldo_financeiro) DESC,
                    codigo_fornecedor
            """
            df_resumo = pd.read_sql(query_resumo, self.conn)

            # Garantir que as colunas sejam float antes de exportar
            for col in ["Saldo Contábil", "Saldo Financeiro", "Diferença (Contábil - Financeiro)"]:
                if col in df_resumo.columns:
                    df_resumo[col] = pd.to_numeric(df_resumo[col], errors="coerce").fillna(0)

            df_resumo.to_excel(writer, sheet_name='Resumo da Conciliação', index=False)

            # Cria aba de Metadados
            metadata_items = [
                'Data e Hora do Processamento',
                'Período de Referência',
                'Total de Fornecedores Processados',
                'Conciliações Conferidas',
                'Conciliações Divergentes',
                'Conciliações Pendentes',
                'Total Financeiro (R$)',
                'Total Contábil (R$)',
                'Saldo Líquido da Conciliação (R$)',
                'Legenda de Status',
                'Tolerância de Diferença'
            ]

            metadata_values = [
                datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                f'{data_inicial} a {data_final}',
                int(stats['total_registros']),
                int(stats['conciliados_ok']),
                int(stats['divergentes']),
                int(stats['pendentes']),
                f"R$ {stats['total_financeiro']:,.2f}",
                f"R$ {stats['total_contabil']:,.2f}",
                f"R$ {stats['diferenca_geral']:,.2f}",
                f"R$ {stats['total_divergencia']:,.2f}",
                'Diferença = Saldo Contábil - Saldo Financeiro',
                'CONFERIDO: Diferença dentro da tolerância (até 3%) | DIVERGENTE: Diferença significativa | PENDENTE: Sem correspondência',
                'Até 3% de discrepância é considerada tolerável'
            ]

            # VERIFICAÇÃO DE COMPRIMENTO - adicione esta validação
            if len(metadata_items) != len(metadata_values):
                logger.error(f"Metadados incompatíveis: {len(metadata_items)} itens vs {len(metadata_values)} valores")
                # Ajusta para ter o mesmo comprimento
                min_length = min(len(metadata_items), len(metadata_values))
                metadata_items = metadata_items[:min_length]
                metadata_values = metadata_values[:min_length]

            metadata = {
                'Item': metadata_items,
                'Valor': metadata_values
            }

            df_metadata = pd.DataFrame(metadata)
            df_metadata.to_excel(writer, sheet_name='Metadados', index=False)
            
            writer.close()
            
            # Aplica estilos ao arquivo gerado
            workbook = openpyxl.load_workbook(output_path)
            
            # Aplica estilos à aba Metadados
            if 'Metadados' in workbook.sheetnames:
                meta_sheet = workbook['Metadados']
                self._apply_metadata_styles(meta_sheet)
            
            # Aplica estilos melhorados à aba Resumo da Conciliação
            if 'Resumo da Conciliação' in workbook.sheetnames:
                resumo_sheet = workbook['Resumo da Conciliação']
                self._apply_enhanced_styles(resumo_sheet, stats)
                
                # Adiciona filtros automáticos
                resumo_sheet.auto_filter.ref = resumo_sheet.dimensions
            
            if 'Resumo da Conciliação' in workbook.sheetnames:
                resumo_sheet = workbook['Resumo da Conciliação']
                self._apply_enhanced_styles(resumo_sheet, stats)
                
                # Adiciona filtros automáticos
                resumo_sheet.auto_filter.ref = resumo_sheet.dimensions
            
            # Aplica estilos básicos às outras abas
            for sheetname in workbook.sheetnames:
                if sheetname not in ['Resumo da Conciliação', 'Metadados']:
                    sheet = workbook[sheetname]
                    self._apply_styles(sheet)
            
            
            # Protege todas as abas (exceto coluna Observações)
            self._protect_sheets(workbook)
            
            workbook.save(output_path)
            
            # Valida o arquivo gerado
            if not self.validate_output(output_path):
                raise ValueError("A validação da planilha gerada falhou")
            
            logger.info(f"Arquivo exportado com sucesso: {output_path}")
            return output_path
        except Exception as e:
            error_msg = f"Erro ao exportar resultados: {e}"
            logger.error(error_msg)
            raise ResultsSaveError(error_msg, caminho=output_path) from e

    def validate_output(self, output_path):
        """
        Valida a estrutura do arquivo Excel gerado e a formatação monetária.
        """
        try:
            wb = openpyxl.load_workbook(output_path)
            
            # Verifica abas obrigatórias
            required_sheets = ['Resumo da Conciliação', 'Títulos a Pagar', 'Balancete', 'Contas x Itens', 'Metadados']
            for sheet in required_sheets:
                if sheet not in wb.sheetnames:
                    raise ValueError(f"Aba '{sheet}' não encontrada no arquivo gerado")
            
            # Verifica formatação monetária na aba Resumo
            resumo = wb['Resumo da Conciliação']
            monetary_columns = ['Valor Financeiro', 'Valor Contábil', 'Diferença']
            
            header = [cell.value for cell in resumo[1] if cell.value is not None]
            
            for col_name in monetary_columns:
                if col_name in header:
                    col_idx = header.index(col_name) + 1
                    # Verifica se pelo menos uma célula tem formatação monetária
                    sample_cell = resumo.cell(row=2, column=col_idx)
                    if sample_cell.value is not None and hasattr(sample_cell, 'number_format'):
                        if 'R$' not in sample_cell.number_format and '#,##0.00' not in sample_cell.number_format:
                            raise ValueError(f"Coluna '{col_name}' não está formatada como moeda brasileira")
            
            return True
            
        except Exception as e:
            error_msg = f"Validação falhou: {e}"
            logger.error(error_msg)
            return False

    def close(self):
        """Fecha a conexão com o banco de dados"""
        if self.conn:
            try:
                self.conn.close()
                logger.info("Conexão com o banco de dados fechada")
            except Exception as e:
                logger.error(f"Erro ao fechar conexão: {e}")

    def __enter__(self):
        """
        Suporte para context manager (with statement).
        
        Returns:
            DatabaseManager: Instância da classe
        """
        self._initialize_database()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """
        Suporte para context manager (with statement).
        
        Args:
            exc_type: Tipo de exceção (se ocorreu)
            exc_val: Valor da exceção
            exc_tb: Traceback da exceção
        """
        self.close()